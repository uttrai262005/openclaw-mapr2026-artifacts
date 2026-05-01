from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

WS = Path(__file__).resolve().parent
STAT_PATH = WS / 'output' / 'thong_ke_dataset.xlsx'


def get_int(v):
    if v is None:
        return None
    try:
        if isinstance(v, str) and v.strip() == '':
            return None
        return int(v)
    except Exception:
        try:
            return int(float(v))
        except Exception:
            return None


def count_bt_dataset(bt: str) -> int:
    base = WS / 'dataset_clean' / bt
    if bt == 'BT3':
        return sum(1 for p in base.iterdir() if p.is_dir())
    # BT1/2/4: files (docx/pdf/txt/md)
    allowed = {'.docx', '.pdf', '.txt', '.md'}
    return sum(1 for p in base.iterdir() if p.is_file() and p.suffix.lower() in allowed)


def main():
    if not STAT_PATH.exists():
        raise SystemExit(f'NOT_FOUND: {STAT_PATH}')

    actual = {bt: count_bt_dataset(bt) for bt in ['BT1', 'BT2', 'BT3', 'BT4']}

    wb = load_workbook(STAT_PATH, data_only=True)
    ws = wb.active

    # Heuristic: find row containing BT1/BT2/BT3/BT4 and take numeric in same row.
    found = {}
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for bt in ['BT1', 'BT2', 'BT3', 'BT4']:
            if bt in [str(v).strip() for v in row_vals if v is not None]:
                # pick first int in row other than bt itself
                nums = [get_int(v) for v in row_vals]
                nums = [n for n in nums if n is not None]
                if nums:
                    found[bt] = nums[0]

    print('actual_counts', actual)
    print('thong_ke_found', found)

    for bt in ['BT1','BT2','BT3','BT4']:
        a = actual[bt]
        s = found.get(bt)
        if s is None:
            print(bt, '-> could_not_parse_from_thong_ke')
        else:
            print(bt, 'actual', a, 'stat', s, 'match', a == s)


if __name__ == '__main__':
    main()
