from __future__ import annotations

import argparse
import importlib.util
import json
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
SCRIPTS = BASE / "scripts"
if str(BASE) not in sys.path:
    sys.path.insert(0, str(BASE))
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from chairman_agent import combine_results
from extract_text import extract_text
from rubric_bt4_parser import parse_rubric_bt4
from rubric_bt1_parser import parse_rubric_bt1
from rubric_bt2_parser import parse_rubric_bt2
from rubric_bt3_parser import parse_rubric_bt3
from schema import GraderResult


def _load_grade_func(grader_dir: Path) -> Callable[..., GraderResult]:
    """Load `grade()` from a grader.py located under a folder that may contain hyphens."""
    grader_py = grader_dir / "grader.py"
    if not grader_py.exists():
        raise FileNotFoundError(grader_py)

    mod_name = f"_gv2_{grader_dir.name.replace('-', '_')}"
    spec = importlib.util.spec_from_file_location(mod_name, grader_py)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load module from {grader_py}")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)  # type: ignore
    if not hasattr(mod, "grade"):
        raise RuntimeError(f"grader missing grade(): {grader_py}")
    return getattr(mod, "grade")


def _ensure_xlsx(path: Path, headers: List[str]) -> None:
    """Ensure output xlsx exists and has at least the requested headers.

    If the file already exists but is missing some headers, append missing
    header cells to the first row and pad existing rows with blanks.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "phase2"
        ws.append(headers)
        wb.save(path)
        return

    wb = load_workbook(path)
    ws = wb.active

    # read current headers (row 1)
    cur = [c.value for c in ws[1]]
    cur = [str(x) if x is not None else "" for x in cur]
    # trim trailing empty header cells
    while cur and (cur[-1] or "").strip() == "":
        cur.pop()

    # append missing headers
    changed = False
    for h in headers:
        if h not in cur:
            cur.append(h)
            changed = True

    if changed:
        # rewrite header row
        for j, h in enumerate(cur, start=1):
            ws.cell(row=1, column=j, value=h)
        # pad existing data rows to new width
        width = len(cur)
        for i in range(2, ws.max_row + 1):
            for j in range(1, width + 1):
                if ws.cell(row=i, column=j).value is None:
                    # keep as None; just ensure cell exists
                    _ = ws.cell(row=i, column=j)
        wb.save(path)


def _append_row(path: Path, row: List[Any]) -> None:
    wb = load_workbook(path)
    ws = wb.active
    width = ws.max_column
    # pad/truncate to match header width
    if len(row) < width:
        row = row + [""] * (width - len(row))
    elif len(row) > width:
        row = row[:width]
    ws.append(row)
    wb.save(path)


def _load_existing_mssv(path: Path) -> List[str]:
    if not path.exists():
        return []
    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, j).value for j in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]
    try:
        mcol = headers.index("MSSV") + 1
    except ValueError:
        return []

    out: List[str] = []
    for i in range(2, ws.max_row + 1):
        v = ws.cell(i, mcol).value
        if v is None:
            continue
        out.append(str(v).split(".")[0].strip())
    return out


def _drop_rows_by_mssv(path: Path, mssv_list: List[str]) -> int:
    if not path.exists() or not mssv_list:
        return 0
    targets = {str(x).strip() for x in mssv_list if str(x).strip()}
    if not targets:
        return 0

    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, j).value for j in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]
    if "MSSV" not in headers:
        return 0
    mcol = headers.index("MSSV") + 1

    deleted = 0
    # delete from bottom to top
    for i in range(ws.max_row, 1, -1):
        v = ws.cell(i, mcol).value
        if v is None:
            continue
        m = str(v).split(".")[0].strip()
        if m in targets:
            ws.delete_rows(i, 1)
            deleted += 1

    if deleted:
        wb.save(path)
    return deleted


def _load_phase1_totals(phase1_xlsx: Optional[str | Path]) -> Dict[str, float]:
    if not phase1_xlsx:
        return {}
    p = Path(phase1_xlsx)
    if not p.exists():
        return {}
    df = pd.read_excel(p)
    # expected columns: MSSV, Tổng(/10)
    mssv_col = None
    total_col = None
    for c in df.columns:
        if str(c).strip().lower() == "mssv":
            mssv_col = c
        if "tổng" in str(c).lower() and "/10" in str(c).lower():
            total_col = c
    if mssv_col is None or total_col is None:
        return {}
    out: Dict[str, float] = {}
    for _, r in df.iterrows():
        m = str(r[mssv_col]).split(".")[0]
        try:
            out[m] = float(r[total_col])
        except Exception:
            continue
    return out


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="Phase 2 multi-agent grading orchestrator (BT4)")
    ap.add_argument("--rubric", required=True, help="Path to rubric_BT4.docx")
    ap.add_argument("--dataset", required=True, help="Folder dataset_clean/BT4")
    ap.add_argument("--limit", type=int, default=None, help="How many first files to grade (default: all)")
    ap.add_argument("--phase1", default="", help="Phase1 XLSX to compare (optional)")
    ap.add_argument("--out", required=True, help="Output XLSX path")
    ap.add_argument("--reset", action="store_true", help="Delete existing output XLSX before running")
    ap.add_argument("--resume", action="store_true", help="Skip MSSV already present in output XLSX")
    ap.add_argument("--drop-mssv", action="append", default=[], help="Delete rows for these MSSV from output before grading (repeatable)")
    args = ap.parse_args(argv)

    rubric_path = str(args.rubric)
    low = rubric_path.lower()
    if "bt1" in low:
        rubric, rubric_raw = parse_rubric_bt1(args.rubric)
        rubric_kind = "BT1"
    elif "bt2" in low:
        rubric, rubric_raw = parse_rubric_bt2(args.rubric)
        rubric_kind = "BT2"
    elif "bt3" in low:
        rubric, rubric_raw = parse_rubric_bt3(args.rubric)
        rubric_kind = "BT3"
    else:
        rubric, rubric_raw = parse_rubric_bt4(args.rubric)
        rubric_kind = "BT4"

    dataset_dir = Path(args.dataset)

    # Support either: flat files dataset_clean/BT*/*.pdf OR per-student folders (BT3 style)
    files = sorted([p for p in dataset_dir.iterdir() if p.is_file() and p.suffix.lower() in {".pdf", ".docx", ".txt", ".png", ".jpg", ".jpeg", ".webp"}])
    items: List[tuple[str, List[Path]]] = []

    if files:
        items = [(p.stem, [p]) for p in files]
    else:
        # folder-per-student
        for sub in sorted([p for p in dataset_dir.iterdir() if p.is_dir()]):
            mssv = sub.name
            docs = sorted([q for q in sub.rglob('*') if q.is_file() and q.suffix.lower() in {'.pdf','.docx','.txt','.png','.jpg','.jpeg','.webp'}])
            # keep even if empty so we can record as missing/zero-score row
            items.append((mssv, docs))

    if args.limit is not None and int(args.limit) > 0:
        items = items[: int(args.limit)]

    out_xlsx = Path(args.out)
    if args.reset and out_xlsx.exists():
        out_xlsx.unlink()

    if rubric_kind == "BT1":
        headers = [
            "MSSV",
            "TC1(/2)",
            "TC2(/2)",
            "TC3(/2)",
            "TC4(/2)",
            "TC5(/2)",
            "Tổng_P2(/10)",
            "Tổng_P1(/10)",
            "Diff(P2-P1)",
            "Veto",
            "Spread",
            "Total_content",
            "Total_structure",
            "Total_language",
            "Nhận xét (chairman)",
            "Ghi chú",
        ]
    elif rubric_kind == "BT2":
        headers = [
            "MSSV",
            "TC1(/1.5)",
            "TC2(/2)",
            "TC3(/2.5)",
            "TC4(/2)",
            "TC5(/2)",
            "Tổng_P2(/10)",
            "Tổng_P1(/10)",
            "Diff(P2-P1)",
            "Veto",
            "Spread",
            "Total_content",
            "Total_structure",
            "Total_language",
            "Nhận xét (chairman)",
            "Ghi chú",
        ]
    elif rubric_kind == "BT3":
        headers = [
            "MSSV",
            "TC1(/1.5)",
            "TC2(/2)",
            "TC3(/3)",
            "TC4(/2)",
            "TC5(/1.5)",
            "Tổng_P2(/10)",
            "Tổng_P1(/10)",
            "Diff(P2-P1)",
            "Veto",
            "Spread",
            "Total_content",
            "Total_structure",
            "Total_language",
            "Nhận xét (chairman)",
            "Ghi chú",
        ]
    else:
        headers = [
            "MSSV",
            "TC1(/3)",
            "TC2(/4)",
            "TC3(/2.5)",
            "TC4(/0.5)",
            "Tổng_P2(/10)",
            "Tổng_P1(/10)",
            "Diff(P2-P1)",
            "Veto",
            "Spread",
            "Total_content",
            "Total_structure",
            "Total_language",
            "Nhận xét (chairman)",
            "Ghi chú",
        ]
    _ensure_xlsx(out_xlsx, headers)

    # optionally drop rows (e.g., known fallback/mistakes)
    dropped = _drop_rows_by_mssv(out_xlsx, args.drop_mssv)

    phase1_map = _load_phase1_totals(args.phase1)

    base = Path(__file__).resolve().parent
    grade_content = _load_grade_func(base / "skills" / "content-grader")
    grade_structure = _load_grade_func(base / "skills" / "structure-grader")
    grade_language = _load_grade_func(base / "skills" / "language-grader")

    diffs: List[Dict[str, Any]] = []

    existing_mssv = set(_load_existing_mssv(out_xlsx)) if (args.resume and out_xlsx.exists()) else set()

    for mssv, paths in items:
        if args.resume and mssv in existing_mssv:
            continue

        # If folder has no submission files, write a zero row (do not call LLM)
        if not paths:
            p1_total = float(phase1_map.get(str(mssv), 0.0)) if phase1_map else 0.0
            diff = (0.0 - float(p1_total)) if p1_total else None
            if rubric_kind in {"BT1", "BT2", "BT3"}:
                row = [
                    int(mssv) if str(mssv).isdigit() else str(mssv),
                    0.0,
                    0.0,
                    0.0,
                    0.0,
                    0.0,
                    0.0,
                    p1_total if p1_total else "",
                    diff if diff is not None else "",
                    "",
                    0.0,
                    0.0,
                    0.0,
                    0.0,
                    "(no files)",
                    "NO_FILES",
                ]
            else:
                row = [
                    int(mssv) if str(mssv).isdigit() else str(mssv),
                    0.0,
                    0.0,
                    0.0,
                    0.0,
                    0.0,
                    p1_total if p1_total else "",
                    diff if diff is not None else "",
                    "",
                    0.0,
                    0.0,
                    0.0,
                    0.0,
                    "(no files)",
                    "NO_FILES",
                ]
            _append_row(out_xlsx, row)
            existing_mssv.add(str(mssv))
            diffs.append({"mssv": mssv, "total_p2": 0.0, "total_p1": p1_total if p1_total else None, "diff": diff, "veto_flag": False, "spread": 0.0})
            continue

        # concatenate texts (BT3 often has 2 CV files)
        parts: List[str] = []
        for path in paths:
            try:
                parts.append(f"\n\n===== FILE: {path.name} =====\n" + extract_text(path))
            except Exception as e:
                parts.append(f"\n\n===== FILE: {path.name} (extract_error) =====\n{e}")
        text = "\n".join(parts)

        from schema import RateLimitError

        with ThreadPoolExecutor(max_workers=3) as ex:
            futs = [
                ex.submit(grade_content, mssv=mssv, submission_text=text, rubric=rubric),
                ex.submit(grade_structure, mssv=mssv, submission_text=text, rubric=rubric),
                ex.submit(grade_language, mssv=mssv, submission_text=text, rubric=rubric),
            ]
            try:
                results = [f.result() for f in futs]
            except RateLimitError as e:
                # stop immediately so user can switch GPT account, do not continue grading
                raise SystemExit(f"RATE_LIMIT_STOP at MSSV={mssv}: {e}")

        final = combine_results(results, rubric=rubric, veto_spread=2.0)
        p2 = final["phase2"]
        p1_total = float(phase1_map.get(str(mssv), 0.0)) if phase1_map else 0.0
        diff = float(p2["total"]) - float(p1_total) if p1_total else None

        any_fallback = any((not getattr(r, "llm_used", True)) for r in results)
        note = "FALLBACK(midpoint)" if any_fallback else ""

        if rubric_kind in {"BT1", "BT2", "BT3"}:
            row = [
                int(mssv) if str(mssv).isdigit() else str(mssv),
                p2["tc1"],
                p2["tc2"],
                p2["tc3"],
                p2["tc4"],
                p2.get("tc5", ""),
                p2["total"],
                p1_total if p1_total else "",
                diff if diff is not None else "",
                "Y" if final["veto_flag"] else "",
                round(float(final["spread"]), 3),
                round(float(final["grader_totals"].get("content", 0.0)), 3),
                round(float(final["grader_totals"].get("structure", 0.0)), 3),
                round(float(final["grader_totals"].get("language", 0.0)), 3),
                p2.get("comment_short", ""),
                note,
            ]
        else:
            row = [
                int(mssv) if str(mssv).isdigit() else str(mssv),
                p2["tc1"],
                p2["tc2"],
                p2["tc3"],
                p2["tc4"],
                p2["total"],
                p1_total if p1_total else "",
                diff if diff is not None else "",
                "Y" if final["veto_flag"] else "",
                round(float(final["spread"]), 3),
                round(float(final["grader_totals"].get("content", 0.0)), 3),
                round(float(final["grader_totals"].get("structure", 0.0)), 3),
                round(float(final["grader_totals"].get("language", 0.0)), 3),
                p2.get("comment_short", ""),
                note,
            ]
        _append_row(out_xlsx, row)
        existing_mssv.add(str(mssv))

        diffs.append(
            {
                "mssv": mssv,
                "total_p2": p2["total"],
                "total_p1": p1_total if p1_total else None,
                "diff": diff,
                "veto_flag": final["veto_flag"],
                "spread": final["spread"],
            }
        )

    report_path = out_xlsx.with_suffix("")
    report_json = Path(str(report_path) + "_diff_report.json")

    # validation: duplicates + missing + fallback markers
    final_mssv = _load_existing_mssv(out_xlsx)
    dupes = sorted({m for m in final_mssv if final_mssv.count(m) > 1})
    if files:
        dataset_mssv = [p.stem for p in sorted([p for p in dataset_dir.iterdir() if p.is_file() and p.suffix.lower() in {".pdf", ".docx", ".txt", ".png", ".jpg", ".jpeg", ".webp"}])]
    else:
        dataset_mssv = [p.name for p in sorted([p for p in dataset_dir.iterdir() if p.is_dir()])]
    missing = sorted(set(dataset_mssv) - set(final_mssv))

    report_json.write_text(
        json.dumps(
            {
                "rubric": rubric_raw,
                "dropped": dropped,
                "processed": len(diffs),
                "diffs": diffs,
                "validation": {"rows": len(final_mssv), "dupes": dupes, "missing": missing},
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
